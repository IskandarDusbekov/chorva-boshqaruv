from datetime import timedelta
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    FinanceCategory,
    FinanceEntry,
    FinanceStatusChoices,
    FinanceTypeChoices,
    MilkPrice,
    MilkRecord,
    Worker,
    WorkerAdvance,
    WorkerJobType,
)
from .selectors import finance_totals_by_category, get_dashboard_overview, get_worker_payroll_summary


HEADER_FILL = PatternFill("solid", fgColor="0F172A")
INCOME_FILL = PatternFill("solid", fgColor="DCFCE7")
EXPENSE_FILL = PatternFill("solid", fgColor="FFE4E6")
INFO_FILL = PatternFill("solid", fgColor="E0E7FF")
MUTED_FILL = PatternFill("solid", fgColor="F8FAFC")
WHITE_FONT = Font(color="FFFFFF", bold=True)
BOLD_FONT = Font(bold=True)


def _money(value):
    return float(Decimal(value or 0))


def _style_header(row):
    for cell in row:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(sheet):
    for column in sheet.columns:
        width = 12
        letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            width = max(width, min(len(value) + 3, 42))
        sheet.column_dimensions[letter].width = width


def _append_title(sheet, title, subtitle):
    sheet.append([title])
    sheet.append([subtitle])
    sheet["A1"].font = Font(size=16, bold=True, color="0F172A")
    sheet["A2"].font = Font(size=11, color="64748B")
    sheet.append([])


def _append_section(sheet, title):
    sheet.append([])
    sheet.append([title])
    sheet.cell(sheet.max_row, 1).font = Font(size=13, bold=True, color="0F172A")
    sheet.cell(sheet.max_row, 1).fill = INFO_FILL


def _append_header(sheet, values):
    sheet.append(values)
    _style_header(sheet[sheet.max_row])


def _style_total_row(sheet):
    for cell in sheet[sheet.max_row]:
        cell.fill = INFO_FILL
        cell.font = BOLD_FONT


def _date_range(date_from, date_to):
    current = date_from
    while current <= date_to:
        yield current
        current += timedelta(days=1)


def _sum_money(entries, *, entry_type=None, currency=None, status=None):
    total = Decimal("0")
    for item in entries:
        if entry_type and item.entry_type != entry_type:
            continue
        if currency and item.currency != currency:
            continue
        if status and item.status != status:
            continue
        total += Decimal(item.amount or 0)
    return total


def _sum_payments(payments, *, currency=None):
    total = Decimal("0")
    for item in payments:
        if currency and item.currency != currency:
            continue
        total += Decimal(item.amount or 0)
    return total


def build_general_report_workbook(*, date_from, date_to):
    workbook = Workbook()
    overview = get_dashboard_overview(date_from, date_to)
    milk_records = list(MilkRecord.objects.filter(record_date__gte=date_from, record_date__lte=date_to).order_by("record_date"))
    finance_entries = list(
        FinanceEntry.objects.filter(entry_date__gte=date_from, entry_date__lte=date_to)
        .select_related("created_by", "related_milk_record")
        .order_by("entry_date", "created_at")
    )
    payments_list = list(
        WorkerAdvance.objects.filter(advance_date__gte=date_from, advance_date__lte=date_to)
        .select_related("worker", "created_by")
        .order_by("advance_date", "created_at")
    )
    workers_list = list(Worker.objects.select_related("job_type").order_by("full_name"))
    payroll_rows = get_worker_payroll_summary()
    milk_prices = list(MilkPrice.objects.order_by("-effective_from", "-created_at"))
    finance_categories = list(FinanceCategory.objects.order_by("entry_type", "name"))
    worker_job_types = list(WorkerJobType.objects.order_by("name"))

    summary = workbook.active
    summary.title = "Dashboard"
    _append_title(summary, "Saytdagi asosiy Dashboard", f"Davr: {date_from} - {date_to}")
    _append_section(summary, "Dashboard kartalari")
    _append_header(summary, ["Ko'rsatkich", "UZS/USD/Litr", "Qo'shimcha", "Izoh"])
    rows = [
        ("Bugungi sut", _money(overview["today_liters"]), "L", f"{_money(overview['growth_percent'])}% kechagiga nisbatan"),
        ("Oxirgi davr sut hajmi", _money(overview["total_liters"]), "L", f"Eng ko'p: {_money(overview['max_daily'])} L | Eng kam: {_money(overview['min_daily'])} L"),
        ("Ferma asosiy hisobi", _money(overview["farm_balance"]["UZS"]), "UZS", f"{_money(overview['farm_balance']['USD'])} USD"),
        ("Kutilayotgan sut puli", _money(overview["pending_milk_totals"]["UZS"]), "UZS", f"{_money(overview['pending_milk_totals']['USD'])} USD"),
        ("Ichki hisob", _money(overview["internal_balance"]["UZS"]), "UZS", f"{_money(overview['internal_balance']['USD'])} USD"),
        ("Tashqi hisob", _money(overview["external_balance"]["UZS"]), "UZS", f"{_money(overview['external_balance']['USD'])} USD"),
        ("Jami kirim", _money(overview["income_totals"]["UZS"]), "UZS", f"{_money(overview['income_totals']['USD'])} USD"),
        ("Jami chiqim", _money(overview["expense_totals"]["UZS"]), "UZS", f"{_money(overview['expense_totals']['USD'])} USD"),
        ("Moliyaviy amallar soni", len(finance_entries), "ta", "Kirim va chiqim yozuvlari"),
        ("Ishchi to'lovlari soni", len(payments_list), "ta", "Avans va ish haqi yozuvlari"),
    ]
    for label, value, unit, note in rows:
        summary.append([label, value, unit, note])
        for cell in summary[summary.max_row]:
            cell.fill = INFO_FILL if label in {"Ichki balans", "Tashqi balans"} else MUTED_FILL
        summary.cell(summary.max_row, 1).font = BOLD_FONT

    _append_section(summary, "Haftalik progress grafigi")
    _append_header(summary, ["Sana", "Litr", "Progress foiz"])
    for item in overview["weekly_chart"]:
        summary.append([item["date"], item["liters"], item["percent"]])

    _append_section(summary, "Dashboarddagi oxirgi sut yozuvlari")
    _append_header(summary, ["Sana", "Ertalab", "Kechki", "Jami", "Narx", "Valyuta"])
    for item in overview["recent_milk_records"]:
        summary.append([item.record_date, _money(item.morning_liters), _money(item.evening_liters), _money(item.total_liters), _money(item.price_per_liter), item.currency])

    _append_section(summary, "Dashboarddagi oxirgi kirim va chiqimlar")
    _append_header(summary, ["Sana", "Tur", "Kategoriya", "Hisob", "Miqdor", "Valyuta"])
    for item in overview["recent_finances"]:
        summary.append([item.entry_date, item.get_entry_type_display(), item.category, item.get_source_display(), _money(item.amount), item.currency])
        fill = INCOME_FILL if item.entry_type == "income" else EXPENSE_FILL
        for cell in summary[summary.max_row]:
            cell.fill = fill

    milk = workbook.create_sheet("Sut boshqaruvi")
    _append_title(milk, "Saytdagi Sut boshqaruvi", f"Davr: {date_from} - {date_to}")
    _append_section(milk, "Sut kartalari")
    _append_header(milk, ["Ko'rsatkich", "UZS/USD/Litr", "Qo'shimcha"])
    active_price = overview["active_milk_price"]
    milk.append(["Aktiv sut narxi", _money(active_price.price_per_liter) if active_price else "", active_price.currency if active_price else "Narx kiritilmagan"])
    milk.append(["Kutilayotgan sut puli UZS", _money(overview["pending_milk_totals"]["UZS"]), "Default hisobda"])
    milk.append(["Kutilayotgan sut puli USD", _money(overview["pending_milk_totals"]["USD"]), "Default hisobda"])
    milk.append(["Davrdagi jami sut", _money(overview["total_liters"]), "L"])

    _append_section(milk, "Sut yozuvlari jadvali")
    _append_header(milk, ["Sana", "Ertalabki litr", "Kechki litr", "Jami litr", "1 litr narxi", "Valyuta", "Kutilayotgan summa", "Kim kiritdi", "Izoh"])
    for item in milk_records:
        milk.append([
            item.record_date,
            _money(item.morning_liters),
            _money(item.evening_liters),
            _money(item.total_liters),
            _money(item.price_per_liter),
            item.currency,
            _money(item.milk_income_amount),
            item.created_by.full_name if item.created_by else "-",
            item.note,
        ])
        milk.cell(milk.max_row, 4).font = BOLD_FONT
        milk.cell(milk.max_row, 7).font = BOLD_FONT
    milk.append(["Jami", sum((item.morning_liters or Decimal("0") for item in milk_records), Decimal("0")), sum((item.evening_liters or Decimal("0") for item in milk_records), Decimal("0")), sum((item.total_liters for item in milk_records), Decimal("0")), "", "", sum((item.milk_income_amount for item in milk_records), Decimal("0")), "", ""])
    _style_total_row(milk)

    _append_section(milk, "Kutilayotgan sut to'lovlari")
    _append_header(milk, ["Sana", "Sut sanasi", "Litr", "Miqdor", "Valyuta", "Holat", "Hisob", "Izoh"])
    pending_milk = [item for item in finance_entries if item.related_milk_record_id and item.status == FinanceStatusChoices.PENDING]
    for item in pending_milk:
        milk.append([
            item.entry_date,
            item.related_milk_record.record_date if item.related_milk_record else "",
            _money(item.related_milk_record.total_liters) if item.related_milk_record else "",
            _money(item.amount),
            item.currency,
            item.get_status_display(),
            item.get_source_display(),
            item.note,
        ])

    _append_section(milk, "Sut narxlari tarixi")
    _append_header(milk, ["Amal qilish sanasi", "Narx", "Valyuta", "Izoh", "Yaratilgan vaqt"])
    for item in milk_prices:
        milk.append([item.effective_from, _money(item.price_per_liter), item.currency, item.note, item.created_at])

    finance = workbook.create_sheet("Moliya")
    _append_title(finance, "Saytdagi Moliya bo'limi", f"Davr: {date_from} - {date_to}")
    _append_section(finance, "Moliya kartalari")
    _append_header(finance, ["Ko'rsatkich", "UZS", "USD", "Izoh"])
    finance_card_rows = [
        ("Jami kirim", overview["income_totals"]["UZS"], overview["income_totals"]["USD"], "Qabul qilingan kirimlar"),
        ("Jami chiqim", overview["expense_totals"]["UZS"], overview["expense_totals"]["USD"], "Qabul qilingan chiqimlar"),
        ("Ichki balans", overview["internal_balance"]["UZS"], overview["internal_balance"]["USD"], "Ichki hisob"),
        ("Tashqi balans", overview["external_balance"]["UZS"], overview["external_balance"]["USD"], "Tashqi hisob"),
    ]
    for label, uzs, usd, note in finance_card_rows:
        finance.append([label, _money(uzs), _money(usd), note])
        _style_total_row(finance)

    _append_section(finance, "Kategoriya bo'yicha rangli jamlanma")
    _append_header(finance, ["Tur", "Kategoriya", "Yozuvlar soni", "UZS", "USD"])
    for row in finance_totals_by_category(finance_entries):
        finance.append([
            "Kirim" if row["entry_type"] == "income" else "Chiqim",
            row["category"],
            row["count"],
            _money(row["UZS"]),
            _money(row["USD"]),
        ])
        fill = INCOME_FILL if row["entry_type"] == "income" else EXPENSE_FILL
        for cell in finance[finance.max_row]:
            cell.fill = fill

    _append_section(finance, "Oxirgi kirim va chiqimlar jadvali")
    _append_header(finance, ["Sana", "Tur", "Kategoriya", "Hisob", "Holat", "Miqdor", "Valyuta", "Kim kiritdi", "Qabul qilingan sana", "Sutga bog'liqmi", "Izoh"])
    for item in finance_entries:
        finance.append([
            item.entry_date,
            item.get_entry_type_display(),
            item.category,
            item.get_source_display(),
            item.get_status_display(),
            _money(item.amount),
            item.currency,
            item.created_by.full_name if item.created_by else "-",
            item.received_at,
            "Ha" if item.related_milk_record_id else "Yo'q",
            item.note,
        ])
        fill = INCOME_FILL if item.entry_type == "income" else EXPENSE_FILL
        for cell in finance[finance.max_row]:
            cell.fill = fill
        finance.cell(finance.max_row, 6).font = BOLD_FONT
    finance.append(["Jami kirim", "", "", "", "", _money(_sum_money(finance_entries, entry_type=FinanceTypeChoices.INCOME, currency="UZS", status=FinanceStatusChoices.CONFIRMED)), "UZS", "", "Qabul qilingan kirim"])
    finance.append(["Jami kirim", "", "", "", "", _money(_sum_money(finance_entries, entry_type=FinanceTypeChoices.INCOME, currency="USD", status=FinanceStatusChoices.CONFIRMED)), "USD", "", "Qabul qilingan kirim"])
    finance.append(["Jami chiqim", "", "", "", "", _money(_sum_money(finance_entries, entry_type=FinanceTypeChoices.EXPENSE, currency="UZS", status=FinanceStatusChoices.CONFIRMED)), "UZS", "", "Qabul qilingan chiqim"])
    finance.append(["Jami chiqim", "", "", "", "", _money(_sum_money(finance_entries, entry_type=FinanceTypeChoices.EXPENSE, currency="USD", status=FinanceStatusChoices.CONFIRMED)), "USD", "", "Qabul qilingan chiqim"])
    _style_total_row(finance)

    _append_section(finance, "Admin paneldagi moliya kategoriyalari")
    _append_header(finance, ["Kategoriya", "Turi", "Faol", "Yaratilgan vaqt"])
    for item in finance_categories:
        finance.append([item.name, item.get_entry_type_display(), "Ha" if item.is_active else "Yo'q", item.created_at])

    workers = workbook.create_sheet("Xodimlar")
    _append_title(workers, "Saytdagi Xodimlarni boshqarish", f"Davr: {date_from} - {date_to}")
    _append_section(workers, "Ishchilar ro'yxati")
    _append_header(workers, ["Ism", "Ish turi", "Boshlagan sana", "Oylik kuni", "Oylik", "Valyuta", "Faol", "Izoh", "Yaratilgan vaqt"])
    for item in workers_list:
        workers.append([
            item.full_name,
            item.job_type.name if item.job_type else item.get_role_display(),
            item.started_at,
            item.payday_day,
            _money(item.monthly_salary),
            item.currency,
            "Ha" if item.is_active else "Yo'q",
            item.note,
            item.created_at,
        ])
        workers.cell(workers.max_row, 5).font = BOLD_FONT

    _append_section(workers, "Ish turlari")
    _append_header(workers, ["Ish turi", "Faol", "Yaratilgan vaqt"])
    for item in worker_job_types:
        workers.append([item.name, "Ha" if item.is_active else "Yo'q", item.created_at])

    _append_section(workers, "Berilgan summalar jadvali")
    _append_header(workers, ["Sana", "Ishchi", "Turi", "Qaysi oy", "Miqdor", "Valyuta", "Kim kiritdi", "Izoh"])
    for item in payments_list:
        workers.append([
            item.advance_date,
            item.worker.full_name,
            item.get_payment_type_display(),
            item.month_reference,
            _money(item.amount),
            item.currency,
            item.created_by.full_name if item.created_by else "-",
            item.note,
        ])
        workers.cell(workers.max_row, 5).font = BOLD_FONT
    workers.append(["Jami", "", "", "", _money(_sum_payments(payments_list, currency="UZS")), "UZS", "", ""])
    _style_total_row(workers)
    workers.append(["Jami", "", "", "", _money(_sum_payments(payments_list, currency="USD")), "USD", "", ""])
    _style_total_row(workers)

    payroll = workbook.create_sheet("Ishchilar hisobi")
    _append_title(payroll, "Saytdagi Ishchilar hisobi", f"Davr: {date_from} - {date_to}")
    _append_section(payroll, "Oylik, avans, qoldiq va oshiqcha")
    _append_header(payroll, ["Ishchi", "Ish turi", "Oylik", "Valyuta", "Shu oy UZS", "Shu oy USD", "Avans UZS", "Avans USD", "Carry", "Qoldiq/Oshiqcha", "Boshlagan sana", "Oylik kuni"])
    for item in payroll_rows:
        payroll.append([
            item["name"],
            item["role"],
            _money(item["salary"]),
            item["currency"],
            _money(item["month_paid_uzs"]),
            _money(item["month_paid_usd"]),
            _money(item["advance_uzs"]),
            _money(item["advance_usd"]),
            _money(item["carry"]),
            _money(item["remaining"]),
            item["started_at"],
            item["payday_day"],
        ])
        payroll.cell(payroll.max_row, 10).font = BOLD_FONT

    _append_section(payroll, "Shu davrdagi ishchi to'lovlari")
    _append_header(payroll, ["Sana", "Ishchi", "Turi", "Qaysi oy", "Miqdor", "Valyuta", "Izoh"])
    for item in payments_list:
        payroll.append([item.advance_date, item.worker.full_name, item.get_payment_type_display(), item.month_reference, _money(item.amount), item.currency, item.note])
        payroll.cell(payroll.max_row, 5).font = BOLD_FONT

    operations = workbook.create_sheet("Barcha amallar")
    _append_title(operations, "Davr mobaynida qilingan barcha amallar", f"Davr: {date_from} - {date_to}")
    _append_header(operations, ["Sana", "Amal turi", "Nomi / Kimga", "Kategoriya", "Miqdor", "Valyuta / Birlik", "Hisob / Holat", "Kim kiritdi", "Izoh"])
    operation_rows = []
    for item in milk_records:
        operation_rows.append((item.record_date, "Sut yozuvi", "Sut", "Ertalab/Kechki", _money(item.total_liters), "L", f"{_money(item.price_per_liter)} {item.currency}", item.created_by.full_name if item.created_by else "-", item.note))
    for item in finance_entries:
        operation_rows.append((item.entry_date, item.get_entry_type_display(), item.category, item.get_source_display(), _money(item.amount), item.currency, item.get_status_display(), item.created_by.full_name if item.created_by else "-", item.note))
    for item in payments_list:
        operation_rows.append((item.advance_date, item.get_payment_type_display(), item.worker.full_name, item.month_reference, _money(item.amount), item.currency, "Ishchi to'lovi", item.created_by.full_name if item.created_by else "-", item.note))
    for row in sorted(operation_rows, key=lambda value: (value[0], value[1], value[2])):
        operations.append(list(row))
        if row[1] == "Kirim":
            fill = INCOME_FILL
        elif row[1] == "Chiqim":
            fill = EXPENSE_FILL
        else:
            fill = MUTED_FILL
        for cell in operations[operations.max_row]:
            cell.fill = fill

    daily = workbook.create_sheet("Kunlik jamlanma")
    _append_title(daily, "Kunlik jamlanma", f"Har bir sana bo'yicha sut, kirim, chiqim va ishchi to'lovi")
    _append_header(daily, ["Sana", "Ertalabki sut", "Kechki sut", "Jami sut", "Kirim UZS", "Kirim USD", "Chiqim UZS", "Chiqim USD", "Ishchi to'lovi UZS", "Ishchi to'lovi USD", "Amallar soni"])
    milk_by_date = {item.record_date: item for item in milk_records}
    for day in _date_range(date_from, date_to):
        day_finance = [item for item in finance_entries if item.entry_date == day and item.status == FinanceStatusChoices.CONFIRMED]
        day_payments = [item for item in payments_list if item.advance_date == day]
        milk_record = milk_by_date.get(day)
        daily.append([
            day,
            _money(milk_record.morning_liters if milk_record else 0),
            _money(milk_record.evening_liters if milk_record else 0),
            _money(milk_record.total_liters if milk_record else 0),
            _money(_sum_money(day_finance, entry_type=FinanceTypeChoices.INCOME, currency="UZS")),
            _money(_sum_money(day_finance, entry_type=FinanceTypeChoices.INCOME, currency="USD")),
            _money(_sum_money(day_finance, entry_type=FinanceTypeChoices.EXPENSE, currency="UZS")),
            _money(_sum_money(day_finance, entry_type=FinanceTypeChoices.EXPENSE, currency="USD")),
            _money(_sum_payments(day_payments, currency="UZS")),
            _money(_sum_payments(day_payments, currency="USD")),
            len(day_finance) + len(day_payments),
        ])
        daily.cell(daily.max_row, 4).font = BOLD_FONT
        daily.cell(daily.max_row, 11).font = BOLD_FONT

    worker_summary = workbook.create_sheet("Xodim jamlanma")
    _append_title(worker_summary, "Ishchilar bo'yicha jamlanma", f"Davr: {date_from} - {date_to}")
    worker_summary.append(["Ishchi", "Avans UZS", "Avans USD", "Ish haqi UZS", "Ish haqi USD", "Jami UZS", "Jami USD", "Yozuvlar soni"])
    _style_header(worker_summary[4])
    grouped_workers = {}
    for item in payments_list:
        bucket = grouped_workers.setdefault(
            item.worker_id,
            {
                "name": item.worker.full_name,
                "advance_uzs": Decimal("0"),
                "advance_usd": Decimal("0"),
                "salary_uzs": Decimal("0"),
                "salary_usd": Decimal("0"),
                "count": 0,
            },
        )
        key = "advance" if item.payment_type == WorkerAdvance.PaymentTypeChoices.ADVANCE else "salary"
        currency_key = "uzs" if item.currency == "UZS" else "usd"
        bucket[f"{key}_{currency_key}"] += Decimal(item.amount or 0)
        bucket["count"] += 1
    for row in grouped_workers.values():
        total_uzs = row["advance_uzs"] + row["salary_uzs"]
        total_usd = row["advance_usd"] + row["salary_usd"]
        worker_summary.append([
            row["name"],
            _money(row["advance_uzs"]),
            _money(row["advance_usd"]),
            _money(row["salary_uzs"]),
            _money(row["salary_usd"]),
            _money(total_uzs),
            _money(total_usd),
            row["count"],
        ])
        worker_summary.cell(worker_summary.max_row, 6).font = BOLD_FONT
        worker_summary.cell(worker_summary.max_row, 7).font = BOLD_FONT

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A5"
        _autosize(sheet)

    return workbook
